"""
Exportação de resultados para Excel.

Transforma os resultados de uma consulta SQL em um arquivo .xlsx.
"""

import os
import tempfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def exportar_excel(dados, nome_arquivo="resultado.xlsx"):
    """
    Exporta os dados para um arquivo Excel temporário.

    Args:
        dados (list): Lista de registros retornados pelo banco.
        nome_arquivo (str): Nome desejado para o arquivo.

    Returns:
        str: Caminho absoluto do arquivo Excel gerado.
    """

    if not dados:
        raise ValueError("Não existem dados para exportar.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Resultado"

    # Os registros do Supabase normalmente são dicionários.
    if isinstance(dados[0], dict):

        colunas = list(dados[0].keys())

        # Cabeçalho
        for coluna, nome_coluna in enumerate(colunas, start=1):
            worksheet.cell(
                row=1,
                column=coluna,
                value=nome_coluna
            )

        # Dados
        for linha, registro in enumerate(dados, start=2):
            for coluna, nome_coluna in enumerate(colunas, start=1):
                worksheet.cell(
                    row=linha,
                    column=coluna,
                    value=registro.get(nome_coluna)
                )

    else:
        raise ValueError(
            "Formato dos dados não suportado para exportação."
        )

    # Ajusta automaticamente a largura das colunas
    for coluna in worksheet.columns:

        maior_tamanho = 0

        for celula in coluna:

            if celula.value is not None:

                tamanho = len(str(celula.value))

                if tamanho > maior_tamanho:
                    maior_tamanho = tamanho

        numero_coluna = coluna[0].column
        letra_coluna = get_column_letter(numero_coluna)

        worksheet.column_dimensions[letra_coluna].width = (
            min(maior_tamanho + 2, 50)
        )

    # Congela o cabeçalho
    worksheet.freeze_panes = "A2"

    # Ativa filtro no cabeçalho
    worksheet.auto_filter.ref = worksheet.dimensions

    # Cria um arquivo temporário
    arquivo_temporario = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        prefix="lista_siab_",
        delete=False
    )

    caminho = arquivo_temporario.name

    arquivo_temporario.close()

    # Salva o Excel no arquivo temporário
    workbook.save(caminho)

    return caminho